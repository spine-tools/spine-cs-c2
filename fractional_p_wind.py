import sys
from openpyxl import load_workbook, Workbook

filepath = sys.argv[1]

wb = load_workbook(filepath, read_only=True, data_only=True)

worksheet = wb["farms"]
	
rows = worksheet.iter_rows()

first_row = next(rows)
header = [x.value for x in first_row]

bidz_column = header.index("bidz")
cap_column = header.index("Pmax")
bus_column = header.index("bus")

zcap = {}
zbus = {}
for row in rows:
    bidz = row[bidz_column].value
    cap = row[cap_column].value
    bus = row[bus_column].value
    zcap.setdefault(bidz, []).append(cap)
    zbus.setdefault(bidz, []).append(bus)

for key, value in zcap.items():
    value_sum = sum(value)
    print(key, value_sum)
    zcap[key] = [v / 7 / 24 / value_sum for v in value]

wb.close()

out_wb = Workbook()
sheet = out_wb.create_sheet("frac_demand")
sheet.append(["bidz", "bus", "frac_demand"])

for bidz, busses in zbus.items():
    for bus, frac_demand in zip(busses, zcap[bidz]):
        sheet.append([bidz, bus, frac_demand])

out_wb.save("frac_demand.xlsx")
out_wb.close()
